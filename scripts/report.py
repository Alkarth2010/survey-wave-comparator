import pandas as pd
from scripts.summary import build_summary_report 

from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

def _color_scale_status_cells(ws):
    """
    Apply cell-based colouring to the 'status' column
    in Scale_Changes sheet.
    """
    headers = [cell.value for cell in ws[1]]
    status_col_idx = headers.index("status") + 1

    fills = {
        "UNCHANGED": PatternFill("solid", fgColor="C6EFCE"),      # light green
        "ADDED": PatternFill("solid", fgColor="D9E1F2"),          # light blue
        "REMOVED": PatternFill("solid", fgColor="F4CCCC"),        # mild red
        "VALUE RECODED": PatternFill("solid", fgColor="FFEB9C"),  # bright yellow
    }

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=status_col_idx)
        if cell.value in fills:
            cell.fill = fills[cell.value]

def _color_variable_status_cells(ws):
    headers = [cell.value for cell in ws[1]]
    status_col = headers.index("status") + 1

    fills = {
        "COMMON": PatternFill("solid", fgColor="C6EFCE"),
        "REVIEW": PatternFill("solid", fgColor="FFEB9C"),
        "VARIABLE NAME MISMATCH": PatternFill("solid", fgColor="FFEB9C"),
        "VARIABLE LABEL REVIEW": PatternFill("solid", fgColor="FFEB9C"),
        "DROPPED": PatternFill("solid", fgColor="F4CCCC"),
        "ADDED": PatternFill("solid", fgColor="D9E1F2"),
    }

    default_fill = PatternFill("solid", fgColor="FCE4D6")

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=status_col)
        cell.fill = fills.get(cell.value, default_fill)


def _add_question_status_formatting(ws):
    """
    Apply conditional formatting to the 'status' column
    in Question_Changes sheet.
    """
    # Find the status column dynamically (safe)
    headers = [cell.value for cell in ws[1]]
    status_col_idx = headers.index("status") + 1
    status_col_letter = ws.cell(row=1, column=status_col_idx).column_letter

    start_row = 2
    end_row = ws.max_row
    col_range = f"{status_col_letter}{start_row}:{status_col_letter}{end_row}"

    # Light green fill for UNCHANGED
    green_fill = PatternFill(
        start_color="C6EFCE",
        end_color="C6EFCE",
        fill_type="solid"
    )

    # Mild yellow fill for others
    yellow_fill = PatternFill(
        start_color="FFF2CC",
        end_color="FFF2CC",
        fill_type="solid"
    )

    ws.conditional_formatting.add(
        col_range,
        FormulaRule(
            formula=[f'{status_col_letter}{start_row}="UNCHANGED"'],
            fill=green_fill
        )
    )

    ws.conditional_formatting.add(
        col_range,
        FormulaRule(
            formula=[f'{status_col_letter}{start_row}<>"UNCHANGED"'],
            fill=yellow_fill
        )
    )


def _make_table(ws, table_name):
    """
    Convert the used range of a worksheet into an Excel Table
    """
    ref = ws.dimensions  # e.g. A1:E120
    table = Table(displayName=table_name, ref=ref)

    style = TableStyleInfo(
        name="TableStyleLight9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False
    )

    table.tableStyleInfo = style
    ws.add_table(table)


def _add_autofilter(ws, df):
    ws.auto_filter.ref = ws.dimensions

def generate_report(outputs):
    output_path = "output/survey_comparison_report.xlsx"

    with pd.ExcelWriter(output_path) as writer:

        # -------------------------
        # REPORT (FIRST SHEET)
        # -------------------------
        report_df = build_summary_report(
            outputs["questions"],
            outputs["variables"],
            outputs["scales"]
        )

        report_df.to_excel(
            writer,
            sheet_name="Report",
            index=False
        )

        # -------------------------
        # Questions
        # -------------------------
        if outputs.get("questions") is not None and not outputs["questions"].empty:
            outputs["questions"].to_excel(
                writer,
                sheet_name="Question_Changes",
                index=False
            )
            ws = writer.sheets["Question_Changes"]
            _make_table(ws, "QuestionChanges")
            
            # Adding colours to the status
            _add_question_status_formatting(ws)

        # -------------------------
        # Variables
        # -------------------------
        if outputs.get("variables") is not None and not outputs["variables"].empty:
            outputs["variables"].to_excel(
                writer,
                sheet_name="Variable_Changes",
                index=False
            )
            ws = writer.sheets["Variable_Changes"]
            _make_table(ws, "VariableChanges")
            
            # Add status highlighting
            _color_variable_status_cells(ws)    

        # -------------------------
        # Scales
        # -------------------------
        if outputs.get("scales") is not None and not outputs["scales"].empty:
            outputs["scales"].to_excel(
                writer,
                sheet_name="Scale_Changes",
                index=False
            )
            ws = writer.sheets["Scale_Changes"]
            _make_table(ws, "ScaleChanges")
            
            # Add status colouring
            _color_scale_status_cells(ws)

        # -------------------------
        # ADD HYPERLINKS (REPORT)
        # -------------------------
        report_ws = writer.sheets["Report"]
        
        for row in range(2, report_ws.max_row + 1):
            metric_cell = report_ws[f"B{row}"]
            metric_text = metric_cell.value

            if metric_text == "Questions added":
                metric_cell.hyperlink = "#'Question_Changes'!A1"
            elif metric_text == "Questions modified / dropped":
                metric_cell.hyperlink = "#'Question_Changes'!A1"
            elif metric_text == "Questions unchanged":
                metric_cell.hyperlink = "#'Question_Changes'!A1"    
            elif metric_text == "Variables added":
                metric_cell.hyperlink = "#'Variable_Changes'!A1"
            elif metric_text == "Variables removed":
                metric_cell.hyperlink = "#'Variable_Changes'!A1"
            elif metric_text == "Variables marked for review":
                metric_cell.hyperlink = "#'Variable_Changes'!A1"
            elif metric_text == "Variables matched (common)":
                metric_cell.hyperlink = "#'Variable_Changes'!A1"    
            elif metric_text == "Variables with scale changes":
                metric_cell.hyperlink = "#'Scale_Changes'!A1"
            elif metric_text == "Variables with recoded values":
                metric_cell.hyperlink = "#'Scale_Changes'!A1"

            if metric_cell.hyperlink:
                metric_cell.style = "Hyperlink"
        
        _make_table(report_ws, "ReportSummary")
    